import openpyxl
from openpyxl import load_workbook
# 临时生成发放：先生成合并单元格表格，在生成信息表格， 最后合并起来，最后调整表格
# num = 1
tag_a = []
ws = openpyxl.Workbook()
wb = ws.active


def merge():
    for j in range(0, 241):  # 241 = 40*6
        if j % 5 == 0:
            first = 'A' + str(j+1)
            second = 'A' + str(j+5)
            wb.merge_cells(first+':'+second)
    ws.save(r"C:\Users\王丽\Desktop\h3cbmcinfo.xlsx")


def bmcinfo():
    with open(r"C:\Users\王丽\Desktop\bmcinfo.txt", mode='r') as f:
        # print(f.read())
        num = 1
        for info in f.readlines():
            inode = -1
            tag_a = []
            if num % 2 != 0:
                info = info.replace('[', '')
                info = info.replace('[', '')
                info = info.replace(']', '')
                info = info.replace('\'', '')
                info = info.replace(',', '')
                info = info.split()
                wb.append(info)
            else:
                info = info.replace('[', '')
                info = info.replace(']', '')
                info = info.replace('\'', '')
                info = info.strip()
                info = info.split(',')
                # print(info)
                # print(type(info))
                for i in info:
                    inode += 1
                    if 'CPU' in i:
                        tag_index = inode
                        end_tag_index = inode - 1
                        tag_a.append(end_tag_index)
                print(tag_a)
                for i in tag_a:
                    try:
                        # print(info[i:tag_a[tag_a.index(i)+1]])
                        wb.append(info[i:tag_a[tag_a.index(i)+1]])
                    except:
                        # print(info[i:])
                        wb.append(info[i:])
            num += 1
    ws.save(r"C:\Users\王丽\Desktop\bmcinfo1.xlsx")


def insert_langchao():  # 批量插入浪潮mac地址，
    ins_arr = []
    num = 1
    ins_wb = load_workbook(r"C:\Users\王丽\Desktop\bmcinfo.xlsx")
    ins_ws = ins_wb.active
    for row in ins_ws['H2:H241']:
        ins_arr.append(row[0].value)
    print(ins_arr)
    for i in range(0, 40):   # 40 = 240/6 6为网卡数包括bond0
        print(ins_arr[num].split()[1])
        num += 6

    for row in ins_ws['I2:I241']:
        ins_arr.append(row[0].value)
    print(ins_arr)
    for i in range(0, 40):
        print(ins_arr[num].split()[1])
        num += 6
    # for i in ins_arr:
    #     if not i:
    #         print(ins_arr.index(i))


def insert_h3c():  # 批量插入H3C mac地址
    ins_arr = []
    num = 2
    ins_wb = load_workbook(r"C:\Users\王丽\Desktop\19bmc.xlsx")
    ins_ws = ins_wb.active
    # for row in ins_ws['A1:A75']:
    #     for cell in row:
    #         print(cell.value)
    #     num += 5
    # print(ins_arr)
    print("15台")
    print('bmc')
    for i in range(0, 15):   # 110/5=22  5 为网卡数
        print(ins_ws[str(num)][0].value)
        num += 5
    num = 2
    print('mac1')
    for i in range(0, 15):   # 110/5=22  5 为网卡数
        print(ins_ws[str(num)][1].value)
        num += 5
    num = 2
    print('mac2')
    for i in range(0, 15):   # 110/5=22  5 为网卡数
        print(ins_ws[str(num)][3].value)
        num += 5
    num = 78
    print("4台")
    print("bmc")
    for i in range(0, 4):   # 110/5=22  5 为网卡数
        print(ins_ws[str(num)][0].value)
        num += 3
    num = 78
    print("mac1")
    for i in range(0, 4):   # 110/5=22  5 为网卡数
        print(ins_ws[str(num)][1].value)
        num += 3
    num = 78
    print("mac2")
    for i in range(0, 4):   # 110/5=22  5 为网卡数
        print(ins_ws[str(num)][3].value)
        num += 3

    # for row in ins_ws['I2:I111']:
    #     ins_arr.append(row[0].value)
    # print(ins_arr)
    # for i in range(0, 22):
    #     print(ins_arr[num].split()[0])
    #     num += 5


def insert_langchao_bmc():
    ins_arr = []
    num = 1
    ins_wb = load_workbook(r"C:\Users\王丽\Desktop\bmcinfo.xlsx")
    ins_ws = ins_wb.active
    for row in ins_ws['D1:D241']:
        ins_arr.append(row[0].value)
    print(ins_arr)
    for i in range(0, 40):  # 40 = 240/6 6为网卡数包括bond0
        print(ins_arr[num].split()[0])
        num += 6

# insert_langchao()
# merge()


insert_h3c()
# bmcinfo()
