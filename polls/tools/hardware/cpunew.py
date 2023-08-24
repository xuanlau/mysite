import requests
# import bs4
from lxml import etree
import myHeaders
import threading
import openpyxl
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl.styles import Font
# import re


def get_html(url, headers):
    " 获取主页html信息 "
    res = requests.get(url=url, headers=headers)
    res.encoding = "gbk"
    return res.text


def get_page_num(html):
    et_html = etree.HTML(html)
    page_numbers = et_html.xpath("//div[@class='sort-box clearfix']/div[@class='small-page']/span/text()")
    return page_numbers


def get_htmlinfo(html, page):
    # 获取页面文字信息
    num = 0
    excel_insert_lable = 0
    # soup = bs4.BeautifulSoup(html, "html.parser")
    et_html = etree.HTML(html)
    cpuinfourl = et_html.xpath("//div[@class='pro-intro']/h3/a/@href")
    wb = openpyxl.Workbook()  # 创建表格对象
    ws = wb.active     # 激活
    fille = PatternFill('solid', fgColor='ffc7ce')  # 设置填充颜色为 橙色
    font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='000000')
    for i in cpuinfourl:
        newurl = "https://detail.zol.com.cn" + i
        new_html = get_html(newurl, headers)
        etnew_html = etree.HTML(new_html)
        cpu_name = etnew_html.xpath("//h1[@class='product-model__name']/text()")  # 获取CPU名称
        # shiyong_info = etnew_html.xpath("//div[@class='section-content']/ul[@class='product-param-item pi-2"
        #                                "8 clearfix']/li/p/text()")  # 台式机CPU爬取方式
        info = etnew_html.xpath("//tbody/*/td[@nowrap='nowrap']/text()")
        infos = []
        for s in info:
            if (num % 2) != 0:
                infos.append(s)
            num = num + 1
        infos.insert(0, cpu_name[0])
        infos = infos[:-2]    # 非详细页面(第一页面)的CPU信息列表
        cpu_detailed_url = etnew_html.xpath("//div[@class='section-content']/a/@href")   # CPU详细信息URL
        cpu_detailed_url = "https://detail.zol.com.cn" + cpu_detailed_url[0]  # 拼接URL
        detailed_html = get_html(cpu_detailed_url, headers)
        detailed_etnew_html = etree.HTML(detailed_html)
        # fenqu_info = detailed_etnew_html.xpath("//tr/td/text()")
        shiyong_info = detailed_etnew_html.xpath("//tr/th/span/text()")   # 详细页面的指标
        if cpu_name[0].find("AMD") != -1:  # AMD信息获取
            cpu_infos = detailed_etnew_html.xpath("//tr/td/span/text()")
            cpu_infos = cpu_infos[0:-3]
        else:  # INTEL信息获取
            cpu_info = detailed_etnew_html.xpath("//tr/td/span/a/text()")  # CPU、制作工艺、线程数量信息
            cpu_info_detail = detailed_etnew_html.xpath("//tr/td/span/text()")   # 详细信息
            for cpu in cpu_info_detail[0:-3]:
                cpu_info.append(cpu)
            cpu_infos = cpu_info
        for j in cpu_infos:
            j = j.strip()  # 去掉换行符
            cpu_name.append(j)  # 拼接完整CPU信息
        cpu_infos_all = cpu_name   # 详细页面的CPU信息列表
        if len(cpu_infos_all) > len(infos):    # 将详细页面CPU信息多于非详细页面的信息，追加到非详细页面的列表中
            cpu_info_num = len(cpu_infos_all) - len(infos)  # 判断哪个更加详细，长度长则更详细
            cpu_info_tmp = cpu_infos_all[-cpu_info_num:]    # 多出来的信息
            for t in cpu_info_tmp:     # 多出来的，追加到非详细页面的列表中
                infos.append(t)
        cpu_infos_all = infos  # 最终得到的CPU信息
        cpu_lable_name = ["CPU_名称"]
        for n in shiyong_info:           # 拼接完整指标
            cpu_lable_name.append(n)     # 最终完整指标
        if excel_insert_lable == 0:
            ws.append(cpu_lable_name)  # 添加excel表格的第一行标题
            for c in range(1, len(cpu_lable_name) + 1):
                ws.cell(row=1, column=c).fill = fille  # 填充背景色
                ws.cell(row=1, column=c).font = font   # 修改字体颜色
                column_char = chr(c+64)
                ws.column_dimensions[column_char].width = 25
            ws.append(cpu_infos_all)  # 写入excel表格中
        else:
            ws.append(cpu_infos_all)  # 写入excel表格中
        excel_insert_lable = excel_insert_lable + 1
    # wb.save(r"H:\lexun\pythonProject1\%s.xlsx" % page)  # 保存表格数据
    wb.save(r"./%s.xlsx" % page)  # 保存表格数据(linux)


class myThread(threading.Thread):
    def __init__(self, html_new, page):
        threading.Thread.__init__(self)
        self.html_new = html_new
        self.page = page

    def run(self):
        print("Starting " + self.name)
        get_htmlinfo(self.html_new, self.page)
        print("Exiting " + self.name)


if __name__ == "__main__":
    # url = "http://jobs.51job.com/beijing-ftp/114540578.html"
    url = "https://zj.zol.com.cn/"
    # 内存信息
    # url1 = "https://detail.zol.com.cn/memory/"
    # cpu信息
    # url1 = "https://detail.zol.com.cn/cpu/"
    # 服务器CPU
    url1 = "https://detail.zol.com.cn/servercpu/"
    headers_choice = myHeaders.get_user_agent()
    headers = {  # 头部信息
        'User-Agent': headers_choice
    }
    cookie = 'ip_ck=4sWG4f/+j7QuNTI3NjU5LjE2NDk0MDAzNDk%3D; Hm_lvt_ae5edc2bc4fc71370807f6187f0a2dd0=165" \
             "6672205; lv=1675323166; vn=4; z_pro_city=s_provice%3Dbeijing%26s_city%3Dbeijing; userPr" \
             "ovinceId=1; userCityId=0; userCountyId=0; userLocationId=1; realLocationId=1; userFidLo" \
             "cationId=1; questionnaire_pv=1675296018; Adshow=2'
    cookie_dict = {i.split("=")[0]:i.split("=")[-1] for i in cookie.split("; ")}
    html = get_html(url1, headers)
    number = get_page_num(html)  # 获取所有页数
    number = number[1].split('/')[1]  # 处理字符串
    # 多线程爬取
    threads = []

    for k in range(1, int(number)-21):
        url_new = url1 + str(k) + '.html'
        html_new = get_html(url_new, headers)
        thread = myThread(html_new, k)
        thread.start()
        threads.append(thread)

    for thread in threads:
        thread.join()
