import os
import time
import openpyxl
from openpyxl import load_workbook, Workbook
import datetime
# from retrying import retry
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service


def h3c():
    # 临时生成发放：先生成合并单元格表格，在生成信息表格， 最后合并起来，最后调整表格`
    a = Workbook()
    if not os.path.exists(r"bmc.xlsx"):
        a.save(r"bmc.xlsx")
    ws = load_workbook(r"bmc.xlsx")
    wb = ws.active
    for ip in range(2, 3):
        read_ip = '192.168.1.' + str(ip)
        # read_ip = 'https://www.baidu.com'
        print(read_ip)
        try:
            # browser.get('https://' + read_ip + '/#login')
            browser.get(read_ip)
            # browser.maximize_window() # 设置窗口最大化
            time.sleep(2)
        except:
            continue
        browser.find_element(By.ID, 'details-button').click()
        browser.find_element(By.ID, 'proceed-link').click()
        time.sleep(4)
        browser.find_element(By.ID, 'login-username').send_keys('admin')
        browser.find_element(By.ID, 'login-password').send_keys('Password@_')
        # @retry(wait_fixed=10, stop_max_attempt_number=1)
        time.sleep(2)
        browser.find_element(By.CLASS_NAME, 'btn-login').click()
        time.sleep(4)
        tancchuang_btn = browser.find_element(By.CLASS_NAME, 'aui_state_highlight')
        # browser.find_element(By.CLASS_NAME, 'aui_state_highlight').click()
        tancchuang_btn.click()
        # alert = browser.switch_to.alert
        # alert.accept()
        time.sleep(3)

        bmc_macs = browser.find_elements(By.XPATH, '//*[@id="mac"]/span')
        bmc_mac = ''
        for b in bmc_macs:
            # print(b.text)
            if '专用' in b.text:
                bmc_mac = b.text.split('：')[1]  # 输出bmc mac地址
                print(bmc_mac)
        # wb.append(bmc_mac.split())
        browser.find_element(By.CLASS_NAME, 'icons-pcie-state').click()
        time.sleep(2)
        nic_click = browser.find_elements(By.CLASS_NAME, 'nicInfo')
        # nic_click = nic.find_elements(By.TAG_NAME, 'span')
        for i in nic_click:
            end_info = []
            i.click()
            time.sleep(1)
            # nic_info1 = browser.find_elements(By.CLASS_NAME, 'col-md-6') # 暂时不增加网卡的详细信息
            # for nic1 in nic_info1:
            #     # print(nic1.text.split()[1])
            #     nic_arr.append(nic1.text.split()[1])
            # # nic_info2 = browser.find_element(By.CLASS_NAME, 'odd')
            # # nic_info3 = browser.find_element(By.CLASS_NAME, 'even')
            end_info.append(bmc_mac)  # 接收bmc mac地址
            nic_info2 = browser.find_element(By.ID, 'nicDatabase').find_element(By.TAG_NAME, 'tbody').text.split()
            for j in nic_info2:
                if ':' in j:  # 如果含有':' , 判定为mac地址，接收
                    end_info.append(j)  # 最后的结果
            # for j in nic_arr:  # 暂时不增加网卡的详细信息
            #     nic_info2.append(j)
            # print(nic_info2)
            wb.append(end_info)
        # for b in browser.find_elements(By.TAG_NAME, 'a'):
        #     if b.text == '用户&安全':
        #         b.click()
        # browser.find_element(By.ID, 'addUser').click()
        # browser.find_element(By.CLASS_NAME, 'h3c-input-s').send_keys('root')
        # browser.find_element(By.ID, 'editUserPwd').send_keys('kingsoft')
        # browser.find_element(By.ID, 'editUserConfirmPwd').send_keys('kingsoft')
        break
    ws.save(r"bmc.xlsx")


option = Options()
# option.page_load_strategy = 'eager'
option.add_experimental_option("detach", True)
prefs = {
    'profile.default_content_setting_values': {
        'images': 2,
        # 'javascript': 2  # 2即为禁用的意思  禁止加载js和图片
    }
}
# option.add_argument(r'C:\Users\11298\AppData\Local\Google\Chrome\User Data')
option.add_experimental_option('excludeSwitches', ['enable-automation'])
option.add_experimental_option('prefs', prefs)
# option.add_argument('headless')  # 不显示浏览器窗口?
option.add_argument('start-maximized')  # 最大化运行（全屏窗口）,不设置，取元素会报错
option.add_argument('disable-infobars')  # 禁用浏览器正在被自动化程序控制的提示
option.add_argument('incognito')  # 隐身模式（无痕模式）
# s = Service(r'C:\Program Files\Google\Chrome\Application\chrome.exe')
# options.addArguments('proxy-bypass-list=*')
browser = webdriver.Chrome(options=option)
start_time = datetime.datetime.now()
h3c()
end_time = datetime.datetime.now()
print(end_time - start_time)
