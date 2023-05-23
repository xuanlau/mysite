import pdfplumber
from openpyxl import workbook, load_workbook
from openpyxl.styles import *
# 自动生成报销市内交通费所需的报销明细表格

pdf_document = r"D:\myfiles\我的文档\报销\20230331\滴滴出行行程报销单.pdf"
bx_mb = r"D:\myfiles\我的文档\报销\报销明细单_乐讯模板.xlsx"
text = ''
row_num = 3
total_money = 0
wb = load_workbook(bx_mb)  # 打开一个已经存在的表格
ws = wb.active

with pdfplumber.open(pdf_document) as pdf_info:
    for page in pdf_info.pages:
        text = text + page.extract_text()

for line in text.split("\n"):
    if '滴滴快车' in line:
        line_info = line.split(' ')
        # 将 line的第3和最后的索引值插入到报销明细表中
        total_money += float(line_info[len(line_info)-1])  # 计算报销总额
        bx_info = [line_info[0], '交通费', '2023-' + line_info[2], '库房办公', line_info[len(line_info)-1], '滴滴打车']
        num = 0
        for cell in ws[str(row_num)]:
            print(num)
            print(bx_info[num])
            cell.value = bx_info[num]
            num += 1
            if num == len(bx_info):
                break
        row_num += 1
align = Alignment(horizontal='center', vertical='center')  # 设置单元格内容居中
print(total_money)
merge_row_num = 29
# ws.unmerge_cells('A' + str(merge_row_num) + ':' + 'E' + str(merge_row_num))
ws.merge_cells('A' + str(merge_row_num) + ':' + 'D' + str(merge_row_num))  # 合并范围内的单元格
ws['A29'] = '合计'
ws['E29'] = total_money
ws['F29'] = '交通方式'
area = ws['A1:F29']  # 设定范围
for are in area:  # 遍历范围内的每行单元格
    for i in are:  # 遍历范围内的每行单元格的每个单元格
        i.border = Border(right=Side(style='thin'), bottom=Side(style='thin'))  # 设置边框为细边框
        i.alignment = align    # 设置单元格内容居中
        i.font = Font(size=10)  # 设置字体大小10
for i in range(1, ws.max_row+1):   # 获取内容所有行行数
    ws.row_dimensions[i].height = 30  # 设置行高为39

wb.save(r'D:\myfiles\我的文档\报销\demaxi.xlsx')  # 另存为
